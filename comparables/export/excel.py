"""Export Excel formate du tableau de comparables (openpyxl). Renvoie des bytes (telechargement)."""
from __future__ import annotations
import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from comparables.models import CompanyRecord
from comparables.config import settings
from comparables.finance.beta import reliable_beta, sample_summary
from comparables.finance.multiples import summary_stats

# (champ, libelle, type, largeur)
DISPLAY = [
    ("name", "Societe", "texte", 24),
    ("ticker", "Ticker", "texte", 10),
    ("country", "Pays", "texte", 13),
    ("sector", "Secteur", "texte", 18),
    ("currency", "Devise", "texte", 7),
    ("market_cap", "Capitalisation", "M", 15),
    ("net_debt", "Dette nette", "M", 13),
    ("enterprise_value", "VE", "M", 15),
    ("beta_source", "Beta (source)", "beta", 12),
    ("index_used", "Indice ref.", "texte", 11),
    ("beta_regression", "Beta (regression)", "beta", 15),
    ("r2", "R2", "r2", 8),
    ("beta_std_err", "Ecart-type beta", "beta", 13),
    ("zero_return_share", "% rdt nuls", "pct", 11),
    ("beta_start", "Debut beta", "texte", 12),
    ("beta_end", "Fin beta", "texte", 11),
    ("gearing", "Gearing net", "pct", 12),
    ("beta_unlevered", "Beta desendette", "beta", 15),
    ("beta_unlevered_adjusted", "Beta desend. ajuste", "beta", 16),
    ("ev_sales", "VE/CA", "mult", 9),
    ("ev_ebitda", "VE/EBITDA", "mult", 10),
    ("ev_ebit", "VE/EBIT", "mult", 9),
    ("pe_trailing", "PER (publie)", "mult", 11),
    ("pe_forward", "PER (estime)", "mult", 11),
    ("pb", "P/B", "mult", 8),
]
STATS_FIELDS = ["beta_source", "beta_regression", "r2", "gearing", "beta_unlevered",
                "beta_unlevered_adjusted",
                "ev_sales", "ev_ebitda", "ev_ebit", "pe_trailing", "pe_forward", "pb"]
# Champs derives de la regression : exclus des stats quand R2 < settings.beta_min_r2.
BETA_QUALITY_FIELDS = ("beta_regression", "beta_unlevered", "beta_unlevered_adjusted")

_FMT = {"M": "# ##0", "beta": "0.00", "r2": "0.00", "pct": "0.0%", "mult": '0.0"x"'}
_LOW_R2_COLOR = "B45309"   # ambre : beta affiche mais exclu des statistiques

# --- Feuille « Synthese » : replique la mise en forme des fichiers maison du cabinet ---
SYNTHESE_ROUGE = "FF0000"        # titres rouges (Roboto Light 16 gras)
SYNTHESE_CYAN = "FF00B0F0"       # fond des en-tetes de tableau (avec alpha FF, parite gabarit)
SYNTHESE_MARINE = "FF1B254C"     # fond de la ligne Mediane (avec alpha FF)
SYNTHESE_POLICE = "Roboto Light"
SYNTHESE_POLICE_BOLD = "Roboto Bold"
# Formats comptables exacts du gabarit maison (2 decimales pour betas/R2, 0 pour montants).
_FMT_COMPTA_2 = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
_FMT_COMPTA_0 = '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-'

BLUME_PENTE = 0.67               # Beta ajuste = 0,67 x beta desendette + 0,33 (convergence vers 1)
BLUME_ORDONNEE = 0.33


def blume_adjusted(beta_unlevered: Optional[float]) -> Optional[float]:
    """Beta ajuste (type Blume) applique par societe : 0,67 x beta desendette + 0,33."""
    if beta_unlevered is None:
        return None
    return BLUME_PENTE * beta_unlevered + BLUME_ORDONNEE


_STAT_LABELS = {"median": "Mediane", "mean": "Moyenne", "min": "Minimum", "max": "Maximum"}


def _low_r2(rec: CompanyRecord) -> bool:
    """Regression trop faible pour exploiter la pente (meme regle que l'ecran)."""
    return (rec.beta_regression is not None
            and reliable_beta(rec.beta_regression, rec.r2, settings.beta_min_r2) is None)


def _stats(records: list[CompanyRecord]) -> dict[str, dict[str, float]]:
    # Delegue a summary_stats (filtre None ET inf/nan) : memes stats que l'ecran.
    # Les champs beta excluent en plus les R2 < beta_min_r2 (pente non exploitable).
    out: dict[str, dict[str, float]] = {}
    for f in STATS_FIELDS:
        if f in BETA_QUALITY_FIELDS:
            values = (reliable_beta(getattr(r, f), r.r2, settings.beta_min_r2)
                      for r in records)
        else:
            values = (getattr(r, f) for r in records)
        s = summary_stats(values)
        if s:
            out[f] = {_STAT_LABELS[k]: v for k, v in s.items()}
    return out


# Couleurs des bordures epaisses de la feuille Synthese (gabarit maison).
SYNTHESE_BORDURE = "FF00B0F0"          # cyan : cadres, separateurs, ligne Mediane
SYNTHESE_BORDURE_VEDETTE = "FF002060"  # bleu fonce : gauche/droite de la cellule vedette


def _thick(left: bool = False, right: bool = False,
           top: bool = False, bottom: bool = False,
           color: str = SYNTHESE_BORDURE,
           lr_color: Optional[str] = None) -> Border:
    """Bordure epaisse cyan ne posant que les cotes demandes (cadres des tableaux).

    `lr_color` surcharge la couleur des cotes gauche/droite (cellule vedette : bleu fonce).
    """
    t = Side(style="thick", color=color)
    tlr = Side(style="thick", color=lr_color or color)
    return Border(left=tlr if left else None, right=tlr if right else None,
                  top=t if top else None, bottom=t if bottom else None)


# Bloc 1 « Donnees sources » : replique la geometrie et les libelles du fichier Capital IQ
# (colonnes B..M). Chaque entree : (libelle, type, getter) ; le type pilote le format.
# F (« Taux d'IS ») et H (« Dette nette ») sont adaptes a nos conventions ; « ($M) » -> « (M) ».
_SYNTH_COLS = [
    ("Sociétés", "texte", lambda r: r.name),
    ("Tickers", "texte", lambda r: r.ticker),
    ("Levered Beta", "beta", lambda r: r.beta_regression),
    # Adj Beta = Levered Beta dans les fichiers Capital IQ (parite visuelle).
    ("Adj Beta", "beta", lambda r: r.beta_regression),
    ("Taux d'IS", "pct", lambda r: settings.tax_rate),
    ("Mkt. Val. Equity (M)", "montant", lambda r: _millions(r.market_cap)),
    ("Dette nette (M)", "montant", lambda r: _millions(r.net_debt)),
    ("Pref Equity (M)", "montant", lambda r: "NA"),           # on ne suit pas les preferentielles
    ("Debt/ Equity", "pct", lambda r: r.gearing),
    ("Pref/ Equity", "pct", lambda r: 0),
    ("Unlevered Beta", "beta", lambda r: r.beta_unlevered),
    ("R2 Correlation", "beta", lambda r: r.r2),
]
_SYNTH_FMT = {"montant": _FMT_COMPTA_0, "beta": _FMT_COMPTA_2, "pct": "0%"}


def _millions(v: Optional[float]) -> Optional[float]:
    return None if v is None else v / 1_000_000.0


def build_synthese_sheet(wb: Workbook, records: list[CompanyRecord],
                         libelle: str = "Échantillon") -> None:
    """Ajoute en premiere position la feuille « Synthese » au format des fichiers maison.

    Deux tableaux : « Donnees sources » (valeurs, colonnes B..M, gabarit Capital IQ) puis
    « Donnees a retenir » (colonnes B..F, formules de reference au bloc 1 + ligne Mediane
    dont la mediane du beta ajuste est mise en exergue) a copier-coller dans le dossier.
    """
    ws = wb.create_sheet("Synthese", 0)
    ws.sheet_view.showGridLines = False

    title_font = Font(name=SYNTHESE_POLICE, size=16, bold=True, color=SYNTHESE_ROUGE)
    head_font = Font(name=SYNTHESE_POLICE_BOLD, size=12, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=SYNTHESE_CYAN)
    txt_font = Font(name=SYNTHESE_POLICE, size=10, color="000000")
    num_font = Font(name=SYNTHESE_POLICE, size=11, color="000000")
    med_font = Font(name=SYNTHESE_POLICE, size=10, bold=True, color="FFFFFF")
    star_font = Font(name=SYNTHESE_POLICE, size=12, bold=True, color="FFFFFF")
    med_fill = PatternFill("solid", fgColor=SYNTHESE_MARINE)
    star_fill = PatternFill("solid", fgColor=SYNTHESE_CYAN)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    head_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # En-tete du bloc 2 (cellule libelle) : aligne a gauche par defaut (horizontal=None), comme le gabarit.
    label_align = Alignment(vertical="center", wrap_text=True)
    # Ligne Mediane : jamais de retour a la ligne (wrap_text=False).
    med_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    txt_align = Alignment(horizontal="left", vertical="center")
    title_align = Alignment(horizontal="left", vertical="center")

    def _title(row: int, col: int, text: str) -> None:
        c = ws.cell(row=row, column=col, value=text)
        c.font = title_font
        c.alignment = title_align
        ws.row_dimensions[row].height = 21

    # Largeurs du gabarit (les autres colonnes restent par defaut ; A et N = marges).
    ws.column_dimensions["A"].width = 9.14
    ws.column_dimensions["B"].width = 60.71
    ws.column_dimensions["C"].width = 20.71
    ws.column_dimensions["D"].width = 20.71
    ws.column_dimensions["E"].width = 20.71
    ws.column_dimensions["F"].width = 20.71
    ws.column_dimensions["N"].width = 9.14

    n = len(records)
    first_col, last_col = 2, 1 + len(_SYNTH_COLS)     # B .. M (12 colonnes)

    # ---- Bloc 1 : Donnees sources ----
    _title(1, 2, "Données sources — extraction NCFcomps")

    head_row = 2
    # En-tetes : top:thick partout, left:thick sur la 1re colonne, right:thick sur CHAQUE
    # cellule (separateurs verticaux du gabarit).
    for j, (label, _typ, _get) in enumerate(_SYNTH_COLS, start=first_col):
        c = ws.cell(row=head_row, column=j, value=label)
        c.font = head_font
        c.fill = head_fill
        c.alignment = head_align
        c.border = _thick(left=(j == first_col), right=True, top=True)
    ws.row_dimensions[head_row].height = 15

    first_data = head_row + 1
    for k, rec in enumerate(records):
        r = first_data + k
        ws.row_dimensions[r].height = 15
        is_last = k == n - 1
        for j, (_label, typ, get) in enumerate(_SYNTH_COLS, start=first_col):
            val = get(rec)
            c = ws.cell(row=r, column=j, value=val)
            c.fill = white_fill
            if typ == "texte":
                c.font = txt_font
                c.alignment = txt_align
            else:
                c.font = num_font
                if isinstance(val, (int, float)):
                    c.number_format = _SYNTH_FMT[typ]
            c.border = _thick(left=(j == first_col), right=(j == last_col), bottom=is_last)
    last_data = first_data + n - 1 if n else head_row

    # ---- Bloc 2 : Donnees a retenir (une seule ligne vide sous le bloc 1) ----
    title2 = last_data + 2
    _title(title2, 2, "Données à retenir --> tableau à mettre dans les rapports")

    # En-tete : libelle echantillon | Beta endette | Beta desendette | Beta ajuste | R²  (colonnes B..F)
    head2 = title2 + 1
    b2_headers = [libelle, "Beta endetté", "Beta désendetté", "Beta ajusté", "R²"]
    b2_first, b2_last = 2, 6                   # B .. F
    for j, label in enumerate(b2_headers, start=b2_first):
        c = ws.cell(row=head2, column=j, value=label)
        c.font = head_font
        c.fill = head_fill
        # Cellule libelle (1re colonne) alignee a gauche par defaut ; les autres centrees.
        c.alignment = label_align if j == b2_first else head_align
        c.border = _thick(left=(j == b2_first), right=True, top=True)
    ws.row_dimensions[head2].height = 15

    # Lignes de donnees : formules de reference au bloc 1 (pas de recopie de valeurs), sans fill.
    # Bloc 1 : D=Levered Beta (endette), L=Unlevered Beta (desendette), M=R². Beta ajuste = Blume
    # en clair sur la cellule Beta desendette de la meme ligne du bloc 2.
    b2_first_data = head2 + 1
    for k in range(n):
        src = first_data + k                   # ligne du bloc 1
        r = b2_first_data + k
        ws.row_dimensions[r].height = 15
        formules = [(2, f"=+B{src}", "texte"), (3, f"=+D{src}", "beta"),
                    (4, f"=+L{src}", "beta"), (5, f"=+D{r}*67%+1*33%", "beta"),
                    (6, f"=+M{src}", "beta")]
        for j, formule, typ in formules:
            c = ws.cell(row=r, column=j, value=formule)
            if typ == "texte":
                c.font = txt_font
                c.alignment = txt_align
            else:
                c.font = num_font
                c.number_format = _FMT_COMPTA_2
            c.border = _thick(left=(j == b2_first), right=(j == b2_last))
    b2_last_data = b2_first_data + n - 1 if n else head2

    # Ligne Mediane : fond marine, texte blanc gras, formules MEDIAN, bordures thick haut/bas.
    # E (mediane du beta ajuste) = cellule vedette : fond cyan, 12pt gras, encadree thick.
    med_row = b2_last_data + 1
    ws.row_dimensions[med_row].height = 17.25
    star_col = 5                               # colonne E : le chiffre a recopier dans le WACC
    for j in range(b2_first, b2_last + 1):
        c = ws.cell(row=med_row, column=j)
        est_vedette = j == star_col
        c.font = star_font if est_vedette else med_font
        c.fill = star_fill if est_vedette else med_fill
        # Aucune cellule de la ligne Mediane ne fait de retour a la ligne (wrap_text=False).
        c.alignment = txt_align if j == b2_first else med_center
        c.border = _thick(left=(j == b2_first or est_vedette),
                          right=(j == b2_last or est_vedette),
                          top=True, bottom=True,
                          lr_color=SYNTHESE_BORDURE_VEDETTE if est_vedette else None)
        if j == b2_first:
            c.value = "Médiane"
            c.number_format = "0.0%"     # parite gabarit (sans effet visible sur du texte)
        elif n:
            col = get_column_letter(j)
            c.value = f"=+MEDIAN({col}{b2_first_data}:{col}{b2_last_data})"
            c.number_format = _FMT_COMPTA_2

    # Note rouge sous la mediane (colonne D).
    _title(med_row + 1, 4, "Données à retenir dans le WACC")


def build_workbook(records: list[CompanyRecord], warning: Optional[str] = None,
                   libelle: str = "Échantillon") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparables"

    blue, grey = "1F3864", "D9E1F2"
    head = Font(bold=True, color="FFFFFF", size=10)
    bold = Font(bold=True, size=10)
    ital = Font(italic=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="thin", color="BFBFBF"))

    ws.cell(row=1, column=1, value="Echantillon de comparables boursiers").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value="Source : Yahoo Finance - montants en millions de la devise indiquee").font = ital
    header_row = 4

    for j, (_f, label, _t, width) in enumerate(DISPLAY, start=1):
        c = ws.cell(row=header_row, column=j, value=label)
        c.font = head
        c.fill = PatternFill("solid", fgColor=blue)
        c.alignment = center
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[header_row].height = 34

    r = header_row
    for rec in records:
        r += 1
        low = _low_r2(rec)
        for j, (field, _label, typ, _w) in enumerate(DISPLAY, start=1):
            v = getattr(rec, field)
            if typ == "M" and v is not None:
                v = v / 1_000_000.0
            cell = ws.cell(row=r, column=j, value=("n.d." if v is None else v))
            cell.border = border
            if v is None:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(color="A6A6A6", size=10)
            elif typ in _FMT:
                cell.number_format = _FMT[typ]
                # R2 trop faible : beta visible mais hors stats -> en ambre (meme regle que l'ecran)
                if low and field in (*BETA_QUALITY_FIELDS, "r2"):
                    cell.font = Font(color=_LOW_R2_COLOR, size=10)
            if typ == "texte":
                cell.alignment = Alignment(horizontal="left")

    stats = _stats(records)
    r += 1
    for label in ("Mediane", "Moyenne", "Minimum", "Maximum"):
        r += 1
        cl = ws.cell(row=r, column=1, value=label)
        cl.font = bold
        cl.fill = PatternFill("solid", fgColor=grey)
        for j, (field, _label, typ, _w) in enumerate(DISPLAY, start=1):
            if j == 1:
                continue
            cell = ws.cell(row=r, column=j)
            cell.fill = PatternFill("solid", fgColor=grey)
            if field in stats and label in stats[field]:
                cell.value = stats[field][label]
                cell.font = bold
                if typ in _FMT:
                    cell.number_format = _FMT[typ]

    # Synthese beta : betas moyens RETENUS (R2 >= seuil), endette / desendette / desendette ajuste
    summary = sample_summary(
        ((rec.beta_regression, rec.r2, rec.beta_unlevered) for rec in records),
        settings.beta_min_r2,
    )
    if summary and summary["mean_levered"] is not None:
        r += 2
        head_cell = ws.cell(row=r, column=1,
                            value=f"Synthese beta : {summary['n_retained']} retenu(s), "
                                  f"{summary['n_excluded_low_r2']} ecarte(s) (R2 < {settings.beta_min_r2:.2f})")
        head_cell.font = bold
        lines = [
            ("Beta endette moyen retenu", summary["mean_levered"]),
            ("Beta desendette moyen retenu", summary["mean_unlevered"]),
            ("Beta desendette ajuste (0,67 x desend. + 0,33)", summary["mean_unlevered_adjusted"]),
        ]
        for label, value in lines:
            r += 1
            ws.cell(row=r, column=1, value=label).font = bold
            cell = ws.cell(row=r, column=2, value=value if value is not None else "n.d.")
            cell.font = bold
            if value is not None:
                cell.number_format = _FMT["beta"]

    r += 2
    freq = {"1mo": "mensuels", "1wk": "hebdomadaires"}.get(settings.beta_frequency, settings.beta_frequency)
    notes = [
        f"Hypothese d'impot (desendettement) : {settings.tax_rate:.0%}.",
        f"Beta (regression) et R2 : rendements {freq} du titre regresses sur son indice "
        f"de reference, sur {settings.beta_period}. R2 entre 0 et 1 = part de variance expliquee.",
        "Beta (desendette) = beta de regression / (1 + (1 - IS) x Dette nette / Capi) (Hamada).",
        "Beta (desendette ajuste) = 0,67 x beta desendette + 0,33 x 1 (convergence vers le "
        "beta de marche, type Blume applique a l'actif economique).",
        f"Betas en ambre : R2 < {settings.beta_min_r2:.2f}, affiches mais exclus des "
        "statistiques et de la synthese (pente non exploitable).",
        "Multiples tels que publies : non retraites (exceptionnels, minoritaires, IFRS 16...).",
        "Devises potentiellement differentes : comparer les montants avec prudence.",
    ]
    if warning:
        notes.insert(0, warning)
    for n in notes:
        cell = ws.cell(row=r, column=1, value="- " + n)
        cell.font = Font(italic=True, size=9, color="595959")
        r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    # Feuille « Synthese » en premiere position (les feuilles existantes restent inchangees).
    build_synthese_sheet(wb, records, libelle=libelle)
    return wb


def build_excel_bytes(records: list[CompanyRecord], warning: Optional[str] = None,
                      libelle: str = "Échantillon") -> bytes:
    buf = io.BytesIO()
    build_workbook(records, warning, libelle=libelle).save(buf)
    return buf.getvalue()
