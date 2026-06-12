"""Export Excel formaté des cessions FR (openpyxl). Renvoie des bytes (téléchargement).

Même langage visuel que l'export comparables (export/excel.py) : en-tête navy, pied de
statistiques grisé, notes méthodologiques. Les médianes reprennent la « règle d'or » du
site : bornes de plausibilité + exclusion des extrêmes (cf. parsing.robust_values).
"""
from __future__ import annotations
import io
import statistics
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from comparables.fr.models import Cession
from comparables.fr.parsing import MULT_EBE_BOUNDS, PCT_CA_BOUNDS, robust_values

# (champ, libellé, type, largeur)
DISPLAY = [
    ("nom", "Societe", "texte", 30),
    ("siren", "SIREN", "texte", 11),
    ("naf", "NAF", "texte", 8),
    ("ville", "Ville", "texte", 16),
    ("departement", "Dept", "texte", 6),
    ("date", "Annonce", "texte", 11),
    ("prix", "Prix de cession (EUR)", "eur", 16),
    ("ca", "CA (EUR)", "eur", 14),
    ("ca_annee", "Exercice", "annee", 9),
    ("ebe", "EBITDA = EBE (EUR)", "eur", 15),
    ("pct_ca", "Prix / CA", "pct", 10),
    ("mult_ebe", "Prix / EBITDA", "mult", 12),
    ("objet_social", "Objet social (RNE)", "texte", 45),
    ("url", "Lien BODACC", "texte", 38),
]

_FMT = {"eur": "# ##0", "pct": "0%", "mult": '0.0"x"', "annee": "0"}


def _median(values: list[float]) -> Optional[float]:
    return statistics.median(values) if values else None


def build_workbook(cessions: list[Cession]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cessions FR"

    blue, grey = "1F3864", "D9E1F2"
    head = Font(bold=True, color="FFFFFF", size=10)
    bold = Font(bold=True, size=10)
    ital = Font(italic=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="thin", color="BFBFBF"))

    ws.cell(row=1, column=1,
            value="Cessions de fonds de commerce - France").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1,
            value="Sources publiques gratuites : BODACC (prix), ratios INPI/BCE et comptes "
                  "INPI (CA, EBE), Sirene (identite)").font = ital
    header_row = 4

    for j, (_f, label, _t, width) in enumerate(DISPLAY, start=1):
        c = ws.cell(row=header_row, column=j, value=label)
        c.font = head
        c.fill = PatternFill("solid", fgColor=blue)
        c.alignment = center
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[header_row].height = 30

    r = header_row
    for ces in cessions:
        r += 1
        for j, (field, _label, typ, _w) in enumerate(DISPLAY, start=1):
            v = getattr(ces, field)
            cell = ws.cell(row=r, column=j, value=("n.d." if v is None else v))
            cell.border = border
            if v is None:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(color="A6A6A6", size=10)
            elif typ in _FMT:
                cell.number_format = _FMT[typ]
            if typ == "texte":
                cell.alignment = Alignment(horizontal="left")

    # Pied : médianes robustes (bornes métier + exclusion des extrêmes, comme l'écran)
    pct_kept, _ = robust_values([c.pct_ca for c in cessions], PCT_CA_BOUNDS)
    ebe_kept, _ = robust_values([c.mult_ebe for c in cessions], MULT_EBE_BOUNDS)
    prix = [c.prix for c in cessions if c.prix is not None]
    stats = {
        "prix": (_median(prix), "eur"),
        "pct_ca": (_median(pct_kept), "pct"),
        "mult_ebe": (_median(ebe_kept), "mult"),
    }
    r += 2
    cl = ws.cell(row=r, column=1, value=f"Mediane robuste ({len(cessions)} cessions)")
    cl.font = bold
    cl.fill = PatternFill("solid", fgColor=grey)
    for j, (field, _label, _typ, _w) in enumerate(DISPLAY, start=1):
        if j == 1:
            continue
        cell = ws.cell(row=r, column=j)
        cell.fill = PatternFill("solid", fgColor=grey)
        if field in stats and stats[field][0] is not None:
            value, typ = stats[field]
            cell.value = value
            cell.font = bold
            cell.number_format = _FMT[typ]

    r += 2
    notes = [
        "Prix : extrait du texte libre des annonces BODACC (best-effort).",
        "EBITDA approche par l'EBE des comptes sociaux (convention Banque de France).",
        "CA / EBE : exercice clos precedant la cession ; absents pour les comptes "
        "confidentiels (art. L232-25, ~45 % des depots).",
        f"Medianes robustes : prix/CA borne a {PCT_CA_BOUNDS[0]:.0%}-{PCT_CA_BOUNDS[1]:.0%}, "
        f"prix/EBITDA a {MULT_EBE_BOUNDS[0]:.1f}x-{MULT_EBE_BOUNDS[1]:.0f}x, extremes "
        "statistiques exclus (z-score modifie sur la MAD).",
        "Ordres de grandeur indicatifs, a croiser avec le jugement d'un analyste.",
    ]
    for n in notes:
        cell = ws.cell(row=r, column=1, value="- " + n)
        cell.font = Font(italic=True, size=9, color="595959")
        r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    return wb


def build_cessions_excel_bytes(cessions: list[Cession]) -> bytes:
    buf = io.BytesIO()
    build_workbook(cessions).save(buf)
    return buf.getvalue()
