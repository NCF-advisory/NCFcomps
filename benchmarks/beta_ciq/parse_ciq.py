"""Lecture des exports S&P Capital IQ (.xlsx) et correspondance ticker CIQ -> Yahoo.

Harnais de benchmark autonome (ne modifie aucun code produit). Parse les 8 fichiers
d'AmeliorationBeta/ avec openpyxl (data_only) et expose une liste normalisee de
societes par fichier, avec le beta CIQ publie et les composants necessaires au
desendettement. La correspondance de tickers reste explicite et traçable : tout
ticker non resolu est consigne (jamais de substitution silencieuse).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl

BASE_DIR = Path(__file__).resolve().parents[2] / "AmeliorationBeta"

# Date de fin de fenetre (= date d'export) deduite du nom de fichier.
FILE_END_DATE: dict[str, date] = {
    "BETA 27032026.xlsx": date(2026, 3, 27),
    "BETA FLO 27042026.xlsx": date(2026, 4, 27),
    "BETA flo 25032026.xlsx": date(2026, 3, 25),
    "Beta 01072026 Liste software.xlsx": date(2026, 7, 1),
    "Beta 29062026 FLO.xlsx": date(2026, 6, 29),
    "Beta Flo 021225.xlsx": date(2025, 12, 2),
    "Beta VFLO 0603.xlsx": date(2026, 3, 6),
    "Beta_ACIER_COSTE_240626.xlsx": date(2026, 6, 24),
}

# Libelle FR d'echantillon (pour le rapport) + industrie Damodaran de repli (si Yahoo
# ne suffit pas). Cf. consignes de mission.
FILE_LABEL: dict[str, str] = {
    "BETA 27032026.xlsx": "Materiaux de construction (US)",
    "BETA FLO 27042026.xlsx": "Agroalimentaire / fromages",
    "BETA flo 25032026.xlsx": "Materiaux de construction (mixte)",
    "Beta 01072026 Liste software.xlsx": "Logiciels",
    "Beta 29062026 FLO.xlsx": "Paiements / fintech",
    "Beta Flo 021225.xlsx": "BTP / construction",
    "Beta VFLO 0603.xlsx": "Services informatiques",
    "Beta_ACIER_COSTE_240626.xlsx": "Acier",
}
FILE_DAMODARAN_FALLBACK: dict[str, str] = {
    "BETA 27032026.xlsx": "Building Materials",
    "BETA flo 25032026.xlsx": "Building Materials",
    "BETA FLO 27042026.xlsx": "Food Processing",
    "Beta 01072026 Liste software.xlsx": "Software (System & Application)",
    "Beta 29062026 FLO.xlsx": "Computer Services",  # paiements : cf. justification rapport
    "Beta Flo 021225.xlsx": "Engineering/Construction",
    "Beta VFLO 0603.xlsx": "Computer Services",
    "Beta_ACIER_COSTE_240626.xlsx": "Steel",
}

# Prefixe bourse CIQ -> suffixe Yahoo. "" = symbole nu (US).
_PREFIX_TO_SUFFIX: dict[str, str] = {
    "ENXTPA": ".PA", "ENXTAM": ".AS", "ENXTBR": ".BR",
    "XTRA": ".DE", "DB": ".DE", "SWX": ".SW", "LSE": ".L",
    "BIT": ".MI", "WBAG": ".VI", "OM": ".ST", "OB": ".OL",
    "HLSE": ".HE", "ISE": ".IR", "BME": ".MC", "NZSE": ".NZ",
    "ASX": ".AX", "TSX": ".TO", "TSE": ".T",
    "NYSE": "", "NASDAQGS": "", "NASDAQGM": "", "NASDAQCM": "",
}

# Corrections explicites (ticker CIQ exact -> ticker Yahoo). Documentees dans la mission.
_TICKER_OVERRIDE: dict[str, str] = {
    "NASDAQGS:FISV": "FI",     # Fiserv a change de symbole
    "ENXTPA:74SW": "74SW.PA",  # Axway / 74Software
    "ENXTPA:ALHYP": "ALHYP.PA",
}

# Cellules "ticker" qui ne sont pas des tickers -> mapping manuel ou non couvert.
_NAME_AS_TICKER: dict[str, Optional[str]] = {
    "Salzgitter AG": "SZG.DE",
    "Network International": None,  # delistee
    "Olympic Steel": None,         # cellule sans ticker exploitable
}


@dataclass
class CiqCompany:
    source_file: str
    name: Optional[str]
    ciq_ticker: Optional[str]          # tel qu'ecrit dans le fichier
    yahoo_ticker: Optional[str]        # mappe (None si non resolu)
    map_note: str = ""                 # raison si non resolu / cas particulier
    levered_beta: Optional[float] = None
    adj_beta: Optional[float] = None
    tax_rate: Optional[float] = None
    mkt_val_equity: Optional[float] = None
    total_debt: Optional[float] = None
    pref_equity: Optional[float] = None
    debt_equity: Optional[float] = None
    pref_equity_ratio: Optional[float] = None
    unlevered_beta: Optional[float] = None
    r2: Optional[float] = None
    std_error: Optional[float] = None


@dataclass
class CiqFile:
    filename: str
    label: str
    end_date: date
    companies: list[CiqCompany] = field(default_factory=list)


def _num(v: object) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "NA", "N/A", "NM", "-"):
            return None
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def map_ticker(ciq_ticker: Optional[str], name: Optional[str]) -> tuple[Optional[str], str]:
    """CIQ -> Yahoo. Renvoie (ticker_yahoo | None, note). Jamais de substitution muette."""
    raw = (ciq_ticker or "").strip()
    # Cellule vide : peut-etre un nom sans ticker (societe delistee).
    if not raw:
        if name and name.strip() in _NAME_AS_TICKER:
            mapped = _NAME_AS_TICKER[name.strip()]
            return mapped, ("map manuel" if mapped else "non couverte (sans ticker)")
        return None, "non couverte (sans ticker)"
    if raw in _TICKER_OVERRIDE:
        return _TICKER_OVERRIDE[raw], "override explicite"
    # Cas ou la cellule "ticker" est en fait un nom de societe.
    if ":" not in raw:
        if raw in _NAME_AS_TICKER:
            mapped = _NAME_AS_TICKER[raw]
            return mapped, ("map manuel" if mapped else "non couverte (delistee)")
        return None, f"format ticker inconnu ({raw!r})"
    prefix, symbol = raw.split(":", 1)
    prefix = prefix.strip().upper()
    symbol = symbol.strip()
    if prefix not in _PREFIX_TO_SUFFIX:
        return None, f"prefixe bourse inconnu ({prefix})"
    suffix = _PREFIX_TO_SUFFIX[prefix]
    # Espace dans le symbole (classe d'action) -> tiret facon Yahoo.
    symbol = symbol.replace(" ", "-")
    return f"{symbol}{suffix}", "map par prefixe"


def _find_header_row(ws) -> Optional[int]:
    """Localise la ligne d'en-tete (celle qui contient 'Levered Beta')."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("levered beta" in c and "unlevered" not in c for c in cells):
            return i
    return None


# Correspondance libelle d'en-tete (normalise) -> champ CiqCompany.
_HEADER_MAP = {
    "societes": "name", "sociétés": "name", "company name": "name",
    "tickers": "ciq_ticker", "ticker": "ciq_ticker",
    "levered beta": "levered_beta", "adj beta": "adj_beta",
    "average tax rate (5 yr)": "tax_rate",
    "mkt. val. equity ($m)": "mkt_val_equity",
    "total debt ($m)": "total_debt",
    "pref equity ($m)": "pref_equity",
    "debt/ equity": "debt_equity", "debt/equity": "debt_equity",
    "pref/ equity": "pref_equity_ratio", "pref/equity": "pref_equity_ratio",
    "unlevered beta": "unlevered_beta",
    "r2 correlation": "r2",
    "std error": "std_error",
}
_TEXT_FIELDS = {"name", "ciq_ticker"}


def parse_file(path: Path) -> CiqFile:
    filename = path.name
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_row = _find_header_row(ws)
    if hdr_row is None:
        wb.close()
        raise ValueError(f"En-tete introuvable dans {filename}")
    header_cells = next(ws.iter_rows(min_row=hdr_row, max_row=hdr_row, values_only=True))
    col_field: dict[int, str] = {}
    for ci, cell in enumerate(header_cells):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key in _HEADER_MAP:
            col_field[ci] = _HEADER_MAP[key]

    companies: list[CiqCompany] = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        # Les fichiers CIQ portent un SECOND bloc (recapitulatif desendette) sous une
        # ligne vide. On s'arrete au premier vide une fois le tableau principal entame.
        if all(c is None for c in row):
            if companies:
                break
            continue
        vals: dict[str, object] = {}
        for ci, fld in col_field.items():
            if ci < len(row):
                vals[fld] = row[ci]
        name = vals.get("name")
        name = str(name).strip() if isinstance(name, str) else (name or None)
        ciq_ticker = vals.get("ciq_ticker")
        ciq_ticker = str(ciq_ticker).strip() if isinstance(ciq_ticker, str) else (
            str(ciq_ticker) if ciq_ticker is not None else None)
        # Ligne totalement vide de contenu utile.
        if not name and not ciq_ticker:
            continue
        yahoo, note = map_ticker(ciq_ticker, name)
        companies.append(CiqCompany(
            source_file=filename, name=name, ciq_ticker=ciq_ticker,
            yahoo_ticker=yahoo, map_note=note,
            levered_beta=_num(vals.get("levered_beta")),
            adj_beta=_num(vals.get("adj_beta")),
            tax_rate=_num(vals.get("tax_rate")),
            mkt_val_equity=_num(vals.get("mkt_val_equity")),
            total_debt=_num(vals.get("total_debt")),
            pref_equity=_num(vals.get("pref_equity")),
            debt_equity=_num(vals.get("debt_equity")),
            pref_equity_ratio=_num(vals.get("pref_equity_ratio")),
            unlevered_beta=_num(vals.get("unlevered_beta")),
            r2=_num(vals.get("r2")),
            std_error=_num(vals.get("std_error")),
        ))
    wb.close()
    return CiqFile(
        filename=filename,
        label=FILE_LABEL.get(filename, filename),
        end_date=FILE_END_DATE[filename],
        companies=companies,
    )


def parse_all() -> list[CiqFile]:
    out = []
    for filename in FILE_END_DATE:
        path = BASE_DIR / filename
        if path.exists():
            out.append(parse_file(path))
    return out


if __name__ == "__main__":
    files = parse_all()
    for cf in files:
        resolved = sum(1 for c in cf.companies if c.yahoo_ticker)
        print(f"\n=== {cf.filename} ({cf.label}) fin={cf.end_date} "
              f"| {len(cf.companies)} societes, {resolved} resolues")
        for c in cf.companies:
            flag = c.yahoo_ticker or f"[NON RESOLU: {c.map_note}]"
            print(f"   {c.ciq_ticker or c.name:28s} -> {flag:16s} "
                  f"betaL={c.levered_beta}")
