"""Tests de l'export Excel des cessions FR (openpyxl, sans réseau)."""
from __future__ import annotations

import io

from openpyxl import load_workbook

from comparables.fr.export import build_cessions_excel_bytes
from comparables.fr.models import Cession


def _cessions() -> list[Cession]:
    return [
        Cession(siren="111222333", nom="BOULANGERIE A", ville="PARIS", departement="75",
                date="2025-01-15", prix=150000.0, ca=200000.0, ca_annee=2024,
                ebe=40000.0, pct_ca=0.75, mult_ebe=3.75, naf="10.71C",
                url="https://bodacc.example/1"),
        Cession(siren="444555666", nom="BOULANGERIE B", prix=90000.0, ca=120000.0,
                ebe=25000.0, pct_ca=0.75, mult_ebe=3.6, naf="10.71C", date="2024-11-02"),
        Cession(siren="777888999", nom="SANS COMPTES", prix=50000.0),   # CA absent -> n.d.
    ]


def test_export_cessions_xlsx():
    data = build_cessions_excel_bytes(_cessions())
    assert data[:2] == b"PK"                            # signature zip/xlsx

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "Cessions FR"
    headers = [c.value for c in ws[4]]
    assert "Societe" in headers and "Prix de cession (EUR)" in headers
    # Lignes de données : 3 cessions à partir de la ligne 5
    assert ws.cell(row=5, column=1).value == "BOULANGERIE A"
    assert ws.cell(row=5, column=7).value == 150000.0
    assert ws.cell(row=7, column=8).value == "n.d."     # CA absent affiché n.d.
    # Pied : médiane robuste présente avec l'effectif
    labels = [ws.cell(row=r, column=1).value for r in range(8, 12)]
    assert any(isinstance(v, str) and v.startswith("Mediane robuste (3") for v in labels)


def test_export_cessions_vide_reste_valide():
    data = build_cessions_excel_bytes([])
    assert data[:2] == b"PK"
