"""Tests offline de l'export Excel (comparables/export/excel.py)."""
from __future__ import annotations
import io
import math

from openpyxl import load_workbook

from comparables.export.excel import (
    SYNTHESE_BORDURE,
    SYNTHESE_BORDURE_VEDETTE,
    SYNTHESE_CYAN,
    SYNTHESE_MARINE,
    SYNTHESE_POLICE,
    blume_adjusted,
    build_excel_bytes,
    _FMT_COMPTA_0,
    _FMT_COMPTA_2,
    _stats,
)
from comparables.models import CompanyRecord


def _wb(records, warning=None, libelle="Échantillon"):
    return load_workbook(io.BytesIO(build_excel_bytes(records, warning, libelle=libelle)))


def _find(ws, text, max_row=40, max_col=14):
    """Coordonnee (row, col) de la premiere cellule contenant `text`."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value == text:
                return r, c
    raise AssertionError(f"introuvable : {text!r}")


def _rgb(cell):
    # Renvoie le code ARGB complet (ex. "FF00B0F0") pour comparaison exacte au gabarit.
    return cell.fill.fgColor.rgb if cell.fill.patternType else None


def _side(cell):
    b = cell.border
    return tuple(s.style if s else None for s in (b.left, b.right, b.top, b.bottom))


def _side_color(cell, cote):
    """Couleur ARGB de la bordure `cote` ('left'|'right'|'top'|'bottom'), ou None."""
    s = getattr(cell.border, cote)
    return s.color.rgb if s and s.color else None


# ---- Feuille « Synthese » (replique la feuille 'Synthese 2' des fichiers maison) ----
# En-tetes bloc 1 attendus (colonnes B..M), constantes en dur (aucune dependance a AmeliorationBeta).
BLOC1_HEADERS = ["Sociétés", "Tickers", "Levered Beta", "Adj Beta", "Taux d'IS",
                 "Mkt. Val. Equity (M)", "Dette nette (M)", "Pref Equity (M)",
                 "Debt/ Equity", "Pref/ Equity", "Unlevered Beta", "R2 Correlation"]


def _synth_records():
    return [
        CompanyRecord(ticker="XTRA:SAP", name="SAP", market_cap=156_452_912_432.0,
                      net_debt=7_862_000_000.0, beta_regression=1.0203, r2=0.2685,
                      gearing=0.0503, beta_unlevered=0.9854),
        CompanyRecord(ticker="NYSE:CRM", name="SalesForce", market_cap=112_304_963_862.0,
                      net_debt=10_594_428_000.0, beta_regression=1.1681, r2=0.2769,
                      gearing=0.0943, beta_unlevered=1.0927),
    ]


def test_synthese_premiere_feuille_gridlines_et_largeurs():
    wb = _wb(_synth_records())
    assert wb.sheetnames[0] == "Synthese"
    assert "Comparables" in wb.sheetnames            # feuilles existantes intactes
    ws = wb["Synthese"]
    assert ws.sheet_view.showGridLines is False
    assert ws.column_dimensions["A"].width == 9.14
    assert ws.column_dimensions["B"].width == 60.71
    assert ws.column_dimensions["C"].width == 20.71
    assert ws.column_dimensions["D"].width == 20.71
    assert ws.column_dimensions["E"].width == 20.71
    assert ws.column_dimensions["F"].width == 20.71


def test_synthese_bloc1_entetes_et_geometrie():
    ws = _wb(_synth_records())["Synthese"]
    title = ws.cell(row=1, column=2)
    assert title.value == "Données sources — extraction NCFcomps"
    assert title.font.name == SYNTHESE_POLICE and title.font.size == 16
    assert title.font.color.rgb[-6:] == "FF0000"
    # En-tetes exactement en B2..M2, fond cyan, texte blanc gras.
    for k, label in enumerate(BLOC1_HEADERS):
        cell = ws.cell(row=2, column=2 + k)
        assert cell.value == label
        assert _rgb(cell) == SYNTHESE_CYAN
        assert cell.font.color.rgb[-6:] == "FFFFFF" and cell.font.bold
    # En-tetes : top:thick partout, right:thick sur CHAQUE cellule (separateurs verticaux),
    # left:thick sur la 1re colonne uniquement.
    assert _side(ws.cell(row=2, column=2)) == ("thick", "thick", "thick", None)   # B2
    assert _side(ws.cell(row=2, column=7)) == (None, "thick", "thick", None)      # colonne interne
    assert _side(ws.cell(row=2, column=13)) == (None, "thick", "thick", None)     # M2
    # Bordures thick en CYAN (meme couleur que le fond des en-tetes).
    assert _side_color(ws.cell(row=2, column=2), "top") == SYNTHESE_BORDURE
    assert _side_color(ws.cell(row=2, column=2), "left") == SYNTHESE_BORDURE


def test_synthese_adj_beta_egale_levered_et_formats():
    ws = _wb(_synth_records())["Synthese"]
    # D = Levered Beta = beta_regression ; E = Adj Beta = MEME valeur.
    assert ws.cell(row=3, column=4).value == 1.0203
    assert ws.cell(row=3, column=5).value == 1.0203
    assert ws.cell(row=3, column=4).number_format == _FMT_COMPTA_2
    # F = Taux d'IS (settings.tax_rate, format %) ; G = Mkt cap en millions (compta 0).
    assert ws.cell(row=3, column=6).number_format == "0%"
    assert ws.cell(row=3, column=7).value == 156_452.912432
    assert ws.cell(row=3, column=7).number_format == _FMT_COMPTA_0
    # I = Pref Equity = "NA" ; K = Pref/Equity = 0 en % ; J = Debt/Equity en %.
    assert ws.cell(row=3, column=9).value == "NA"
    assert ws.cell(row=3, column=11).value == 0
    assert ws.cell(row=3, column=10).number_format == "0%"


def test_synthese_bloc1_bordure_bottom_derniere_ligne():
    ws = _wb(_synth_records())["Synthese"]
    # 2 societes -> derniere ligne de donnees = 4 ; bottom thick sur toutes les colonnes B..M.
    for col in range(2, 14):
        assert _side(ws.cell(row=4, column=col))[3] == "thick"
    # La ligne 3 (non derniere) n'a pas de bottom.
    assert _side(ws.cell(row=3, column=7))[3] is None
    # Cadre du bloc 1 en cyan (bottom de la derniere ligne, left sur la colonne B).
    assert _side_color(ws.cell(row=4, column=2), "bottom") == SYNTHESE_BORDURE
    assert _side_color(ws.cell(row=4, column=2), "left") == SYNTHESE_BORDURE


def test_synthese_une_seule_ligne_vide_entre_blocs():
    ws = _wb(_synth_records())["Synthese"]
    # Bloc 1 : donnees en 3-4 ; ligne 5 vide ; titre bloc 2 en 6.
    assert all(ws.cell(row=5, column=c).value is None for c in range(2, 14))
    assert ws.cell(row=6, column=2).value == "Données à retenir --> tableau à mettre dans les rapports"


def test_synthese_libelle_echantillon_propage():
    ws = _wb(_synth_records(), libelle="Logiciels")["Synthese"]
    # 1re cellule d'en-tete du bloc 2 (colonne B) = libelle.
    r, c = _find(ws, "Logiciels")
    assert c == 2
    # Cellule libelle alignee a gauche par defaut (horizontal=None) ; les autres centrees.
    assert ws.cell(row=r, column=c).alignment.horizontal is None
    assert ws.cell(row=r, column=c + 1).alignment.horizontal == "center"
    # Les 4 autres en-tetes du bloc 2.
    assert [ws.cell(row=r, column=c + i).value for i in range(1, 5)] == \
        ["Beta endetté", "Beta désendetté", "Beta ajusté", "R²"]
    # En-tetes du bloc 2 : right:thick sur chaque cellule (separateurs), left:thick sur B.
    assert _side(ws.cell(row=r, column=c)) == ("thick", "thick", "thick", None)
    assert _side(ws.cell(row=r, column=c + 2)) == (None, "thick", "thick", None)


def test_synthese_bloc2_formules_reference_et_blume():
    ws = _wb(_synth_records())["Synthese"]
    r, c = _find(ws, "Échantillon")                  # en-tete bloc 2 (defaut)
    first = r + 1
    # Societe / Beta endette (=+D) / Beta desendette (=+L) / R² (=+M) referencent le bloc 1 (ligne 3).
    assert ws.cell(row=first, column=c).value == "=+B3"
    assert ws.cell(row=first, column=c + 1).value == "=+D3"
    assert ws.cell(row=first, column=c + 2).value == "=+L3"
    assert ws.cell(row=first, column=c + 4).value == "=+M3"
    # Beta ajuste = formule Blume en clair sur la cellule Beta desendette de la meme ligne (D{first}).
    assert ws.cell(row=first, column=c + 3).value == f"=+D{first}*67%+1*33%"
    # Cellules du bloc 2 sans remplissage.
    assert ws.cell(row=first, column=c + 1).fill.patternType is None
    # Bordures : left thick sur B, right thick sur F seulement.
    assert _side(ws.cell(row=first, column=c)) == ("thick", None, None, None)
    assert _side(ws.cell(row=first, column=c + 4)) == (None, "thick", None, None)
    # Helper Blume conserve (utilise ailleurs / tests).
    assert blume_adjusted(0.90) == 0.67 * 0.90 + 0.33
    assert blume_adjusted(None) is None


def test_synthese_ligne_mediane_et_cellule_vedette():
    ws = _wb(_synth_records())["Synthese"]
    r, c = _find(ws, "Médiane")
    med = ws.cell(row=r, column=c)                   # colonne B
    assert _rgb(med) == SYNTHESE_MARINE
    assert med.font.name == SYNTHESE_POLICE and med.font.bold
    assert med.font.color.rgb[-6:] == "FFFFFF"
    # Colonnes C/D/F : marine, formule =+MEDIAN(, format comptable.
    cD = ws.cell(row=r, column=c + 2)
    assert cD.value.startswith("=+MEDIAN(") and cD.number_format == _FMT_COMPTA_2
    assert _rgb(cD) == SYNTHESE_MARINE
    # Aucune cellule de la ligne Mediane ne fait de retour a la ligne.
    assert med.alignment.wrap_text in (None, False)
    assert cD.alignment.wrap_text in (None, False)
    # Cellule vedette E = mediane du Beta ajuste : cyan, 12pt, encadree thick des 4 cotes.
    star = ws.cell(row=r, column=c + 3)
    assert star.value.startswith("=+MEDIAN(")
    assert _rgb(star) == SYNTHESE_CYAN
    assert star.font.size == 12 and star.font.bold
    assert _side(star) == ("thick", "thick", "thick", "thick")
    # Vedette : gauche/droite en bleu fonce, haut/bas en cyan.
    assert _side_color(star, "left") == SYNTHESE_BORDURE_VEDETTE
    assert _side_color(star, "right") == SYNTHESE_BORDURE_VEDETTE
    assert _side_color(star, "top") == SYNTHESE_BORDURE
    assert _side_color(star, "bottom") == SYNTHESE_BORDURE
    # Note WACC juste dessous en colonne D.
    assert ws.cell(row=r + 1, column=4).value == "Données à retenir dans le WACC"


def test_stats_filtrent_inf_et_nan():
    records = [
        CompanyRecord(ticker="A", ev_ebitda=8.0),
        CompanyRecord(ticker="B", ev_ebitda=12.0),
        CompanyRecord(ticker="C", ev_ebitda=math.inf),
        CompanyRecord(ticker="D", ev_ebitda=math.nan),
        CompanyRecord(ticker="E", ev_ebitda=None),
    ]
    s = _stats(records)["ev_ebitda"]
    assert s == {"Mediane": 10.0, "Moyenne": 10.0, "Minimum": 8.0, "Maximum": 12.0}


def test_stats_champ_sans_valeur_finie_absent():
    records = [CompanyRecord(ticker="A", pb=math.inf), CompanyRecord(ticker="B")]
    assert "pb" not in _stats(records)


def test_workbook_montants_en_millions_et_nd():
    records = [CompanyRecord(ticker="WMS", name="ADS", market_cap=2_500_000_000.0)]
    ws = _wb(records)["Comparables"]
    # Ligne de donnees = header_row(4) + 1 ; market_cap est la colonne 6 (cf. DISPLAY).
    assert ws.cell(row=5, column=6).value == 2_500.0
    # Champ absent -> "n.d." (net_debt, colonne 7).
    assert ws.cell(row=5, column=7).value == "n.d."


def test_workbook_ligne_mediane_filtre_inf():
    # r2 >= seuil requis depuis la regle qualite : un beta sans R2 est exclu des stats.
    records = [
        CompanyRecord(ticker="A", beta_regression=1.0, r2=0.5),
        CompanyRecord(ticker="B", beta_regression=2.0, r2=0.5),
        CompanyRecord(ticker="C", beta_regression=math.inf, r2=0.5),
    ]
    ws = _wb(records)["Comparables"]
    # Stats : ligne vide puis Mediane/Moyenne/Min/Max apres les 3 lignes de donnees.
    labels = {ws.cell(row=r, column=1).value: r for r in range(8, 13)}
    r_med = labels["Mediane"]
    # beta_regression est la colonne 11 (cf. DISPLAY).
    assert ws.cell(row=r_med, column=11).value == 1.5
    assert ws.cell(row=labels["Maximum"], column=11).value == 2.0


def test_stats_beta_excluent_faible_r2():
    """Beta avec R2 < seuil : affiche en ambre mais hors mediane/moyenne (regle ecran)."""
    records = [
        CompanyRecord(ticker="A", beta_regression=1.0, r2=0.40, beta_unlevered=0.8),
        CompanyRecord(ticker="B", beta_regression=2.0, r2=0.30, beta_unlevered=1.6),
        CompanyRecord(ticker="C", beta_regression=9.0, r2=0.05, beta_unlevered=8.0),
    ]
    s = _stats(records)
    assert s["beta_regression"]["Mediane"] == 1.5            # sans le 9.0
    assert abs(s["beta_unlevered"]["Moyenne"] - 1.2) < 1e-9  # sans le 8.0
    # le R2 lui-meme decrit tout l'echantillon
    assert s["r2"]["Minimum"] == 0.05


def test_workbook_colonnes_beta_fiabilite():
    """Nouvelles colonnes de fiabilite : ecart-type, % rdt nuls, debut/fin de la fenetre."""
    records = [CompanyRecord(ticker="A", beta_regression=1.1, r2=0.5,
                             beta_std_err=0.2, zero_return_share=0.25,
                             beta_start="2021-03-31", beta_end="2026-02-28")]
    ws = _wb(records)["Comparables"]
    header = {ws.cell(row=4, column=j).value: j for j in range(1, 40)
              if ws.cell(row=4, column=j).value}
    for label in ("Ecart-type beta", "% rdt nuls", "Debut beta", "Fin beta"):
        assert label in header
    assert ws.cell(row=5, column=header["Ecart-type beta"]).value == 0.2
    assert ws.cell(row=5, column=header["% rdt nuls"]).value == 0.25
    assert ws.cell(row=5, column=header["Debut beta"]).value == "2021-03-31"
    assert ws.cell(row=5, column=header["Fin beta"]).value == "2026-02-28"


def test_workbook_synthese_beta_presente():
    records = [
        CompanyRecord(ticker="A", beta_regression=1.0, r2=0.40, beta_unlevered=0.8),
        CompanyRecord(ticker="B", beta_regression=2.0, r2=0.30, beta_unlevered=1.6),
    ]
    ws = _wb(records)["Comparables"]
    cells = [(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value)
             for r in range(1, 30)]
    labels = {label: value for label, value in cells if label}
    assert any(str(k).startswith("Synthese beta") for k in labels)
    assert labels["Beta endette moyen retenu"] == 1.5
    assert abs(labels["Beta desendette moyen retenu"] - 1.2) < 1e-9   # (0.8 + 1.6) / 2
    # Ajustement de l'analyste : 0,67 x desendette moyen + 0,33.
    assert abs(labels["Beta desendette ajuste (0,67 x desend. + 0,33)"] - (0.67 * 1.2 + 0.33)) < 1e-9
